# TODO: Подгрузку работников нужно проводить по открытию и по нажатию кнопки обновить(добавить в угол)

import os
import sys
from typing import Type

from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QIntValidator, QIcon, QColor
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QStyledItemDelegate, QLineEdit, \
    QTableWidgetItem, QListWidgetItem, QAbstractItemView, QPushButton, QWidget, QVBoxLayout
from loguru import logger
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.workbook import Workbook
from qframelesswindow import AcrylicWindow, TitleBar, StandardTitleBar

from app.drivers.attendances import AttendancesDataDriver
from app.models import Employee, MotivationProgram, Department
from app.models.control_models import delete_motivation_program, get_current_roles_by_department_code, thresholds_clear, \
    get_employees_by_motivation_program_id
from app.payslip_report.payslip_report import ReportGenerator
from app.screens.edit_employees_window import EditEmployeesWindow
from app.ui.controllers.table_controller import ThresholdsTableController
from app.ui.main_window_ui import Ui_MainWindow
from app.helpers.helpers import get_icon_from_svg, set_departments, get_department_code
from app.db import get_session
from app.ui.styles import CONFIRM_DIALOG_STYLE, WARNING_DIALOG_STYLE
from app.iiko_business_api.employees import update_employees_from_api

with get_session() as session:
    EMPLOYEES_INIT = session.query(Employee).all()
    ROLES_INIT = session.query(MotivationProgram).all()
    DEPARTMENTS_INIT = session.query(Department).all()


class NumericDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        validator = QIntValidator(0, 1000000, editor)  # Измените диапазон по мере необходимости
        editor.setValidator(validator)
        return editor


class CustomTitleBar(TitleBar):
    """ Custom title bar """

    def __init__(self, parent):
        super().__init__(parent)

        # customize the style of title bar button
        self.minBtn.setHoverColor(Qt.white)
        self.minBtn.setHoverBackgroundColor(QColor(0, 100, 182))
        self.minBtn.setPressedColor(Qt.white)
        self.minBtn.setPressedBackgroundColor(QColor(54, 57, 65))

        # use qss to customize title bar button
        self.maxBtn.setStyleSheet("""
            TitleBarButton {
                qproperty-normalColor: black;
                qproperty-normalBackgroundColor: transparent;
                qproperty-hoverColor: white;
                qproperty-hoverBackgroundColor: rgb(0, 100, 182);
                qproperty-pressedColor: white;
                qproperty-pressedBackgroundColor: rgb(54, 57, 65);
            }
        """)


class SalaryReader(AcrylicWindow):
    def __init__(self, parent=None):
        super().__init__(parent=parent)

        # Создаем внутренний QMainWindow для работы с UI из Designer
        self.main_widget = QMainWindow()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self.main_widget)  # Устанавливаем UI в QMainWindow

        # Добавляем QMainWindow в FramelessWindow
        self.setCentralWidget(self.main_widget)

        # Включить стандартный TitleBar с кнопками
        self.setTitleBar(StandardTitleBar(self.main_widget))
        self.setWindowIcon(QIcon("app/ui/icons/export-icon.png"))
        self.setWindowTitle("Kult Salary Reader")
        # Устанавливаем стиль для заголовка окна и цвет текста заголовка
        self.titleBar.setStyleSheet("font-size: 14px; font-weight: bold; color: white;")
        self.titleBar.raise_()

        self.DEBUG = False
        self.salary_table_controller = AttendancesDataDriver(self.ui.salar_table)
        # Передаем ссылку на объект AttendancesDataDriver для возможности использования методов
        self.payslip_generator = ReportGenerator(self.salary_table_controller)

        self.threshold_table_controller = ThresholdsTableController(self.ui.table_motivate_settings)

        # Выпадающий список выбора отдела
        set_departments(self.ui.department, DEPARTMENTS_INIT)
        self.ui.department.currentIndexChanged.connect(  # При выборе отдела выводим список его программ
            self.set_current_roles
        )

        # Программы мотивации(Список)
        self.set_current_roles()  # На старте выводим список программ выбранного отдела
        self.ui.roles_list.currentItemChanged.connect(  # Заполняем таблицу при выборе роли
            self.fill_role_settings_table
        )
        self.ui.roles_list.itemChanged.connect(self.update_role_in_db)
        self.ui.roles_add_button.clicked.connect(self.add_role)
        self.ui.roles_delete.clicked.connect(self.delete_role)

        # Таблица настройки мотивации
        self.ui.table_motivate_settings.setColumnCount(3)
        self.ui.table_motivate_settings.setHorizontalHeaderLabels(
            ["Выручка в руб.", "Сумма вознаграждения(руб.)", "no_role"]
        )
        # Скрываем третий столбец, он для технических нужд:
        self.ui.table_motivate_settings.setColumnHidden(2, True)
        # Выбираем первую роль в списке и заполняем таблицу на открытии
        self.ui.roles_list.setCurrentRow(0)
        self.fill_role_settings_table()

        # Подключаем кнопки добавления строк в таблицу настройки мотивации
        self.ui.button_add_threshhold.clicked.connect(self.add_row)
        self.ui.button_delete_threshold.clicked.connect(self.remove_selected_row)

        # Устанавливаем иконки для кнопки добавления строки в таблицу порогов мотивации
        icon = get_icon_from_svg("app/ui/icons/add.svg")
        self.ui.button_add_threshhold.setText("")
        self.ui.button_add_threshhold.setIcon(icon)
        # Устанавливаем иконку для кнопки удаления строки из таблицы порогов мотивации
        icon = get_icon_from_svg("app/ui/icons/delete.svg")
        self.ui.button_delete_threshold.setText("")
        self.ui.button_delete_threshold.setIcon(icon)

        self.ui.button_settings_save.clicked.connect(lambda: self.save_table_data_to_db(on_save_button_push=True))

        # Таблица сотрудников привязанных к роли
        self.ui.employees_table.setColumnCount(4)
        self.ui.employees_table.setHorizontalHeaderLabels(["ФИО", "Должность", "Отдел", "Табельный"])
        self.ui.employees_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        current_role = self.ui.roles_list.currentItem()
        if current_role:
            current_role_data = current_role.data(Qt.ItemDataRole.UserRole)
            self.fill_employees_table(current_role_data['role_id'])

        self.ui.button_employees_edit.clicked.connect(self.open_edit_employees_window)

        # Сводная таблица зарплат
        self.ui.date_from.setDate(QDate.currentDate())
        self.ui.date_to.setDate(QDate.currentDate())
        self.ui.salar_table.setColumnCount(15)
        self.ui.salar_table.setHorizontalHeaderLabels(
            ["ФИО",
             "Полное имя",
             "Сумма ЗП\nc 1 по 15",
             "Кол-во полных смен\nс 1 по 15",
             "Кол-во неполных смен\nс 1 по 15",
             "Сумма ЗП\n c 16 до конца месяца",
             "Кол-во полных смен\n c 16 до конца месяца",
             "Кол-во неполных смен\n c 16 до конца месяца",
             "Сумма ЗП\n за весь месяц",
             "Кол-во полных смен\n за весь месяц",
             "Кол-во неполных смен\n за весь месяц",
             "Роль", "Отдел", "Табельный", "id"]
        )
        self.ui.salar_table.setSortingEnabled(True)

        self.ui.refresh_salary.clicked.connect(self.update_and_render_salary_table)

        excel_icon = QIcon('app/ui/icons/excel.svg')
        self.excel_button = QPushButton(icon=excel_icon, parent=self.ui.salar_table)
        self.excel_button.setGeometry(1, 1, 23, 23)
        self.excel_button.clicked.connect(self.export_to_excel)

        # Печать ведомостей в pdf по 4 таблицы(сотрудника) на листе
        self.ui.button_payslip_report.clicked.connect(self.payslip_report_callback)

    def setCentralWidget(self, widget: QWidget):
        # Переопределяем метод, чтобы добавить виджет в FramelessWindow
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(widget)
        self.setLayout(layout)

    # TODO: Применить ко всем ошибкам, которые возникают при работе программы
    def show_error_message(self, message, title=None) -> None:
        """
        Функция для вывода сообщения об ошибке в новом окне
        :param title: Заголовок сообщения
        :param message: Сообщение об ошибке
        """
        if title is None:
            title = "Ошибка"

        msg = QMessageBox(self)
        msg.setStyleSheet(WARNING_DIALOG_STYLE)
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setWindowTitle(title)
        msg.exec()

    def update_and_render_salary_table(self):
        """
        Обновление данных в таблице зарплат
        """
        current_department_code = get_department_code(self.ui.department)
        date_from = self.ui.date_from.date().toPython()
        date_to = self.ui.date_to.date().toPython()

        self.salary_table_controller.update_data(
            department_code=current_department_code, date_from=date_from, date_to=date_to)
        try:
            self.salary_table_controller.render_general_table()
        except Exception as e:
            logger.error(f"Ошибка при обновлении данных таблицы зарплат: {e}")
            self.show_error_message(title="Ошибка при обновлении данных таблицы зарплат", message=str(e))

    def set_current_roles(self):
        """
        Функция колбэк для заполнения списка ролей.
        Вызывается при изменении индекса выбранного элемента списка отделов.(Выборе отдела, отличного от текущего)
        Также применяется когда нужно заполнить список ролей(программ мотивации)
        в соответствии с выбранным в выпадающем списке отделом.
        """
        # Отчищаем список ролей
        self.ui.roles_list.clear()
        # Получаем код выбранного отдела
        department_code = get_department_code(self.ui.department)
        # Обновляем данные о работниках из iiko
        # TODO: Получать при запуске, после обновлять только по кнопке
        with get_session() as session:
            try:
                logger.debug(f"Обновляем сотрудников для {department_code} (До if)")
                if not self.DEBUG:
                    logger.debug(f"Обновляем сотрудников для {department_code} (После if)")
                    update_employees_from_api(session=session)
                    session.commit()  # Сохраняем изменения, если всё успешно
                    print(f"[set_current_roles] Сотрудники для отдела {department_code} обновлены")
            except Exception as e:
                session.rollback()  # Откатываем изменения в случае ошибки
                print(f"[set_current_roles] Ошибка при обновлении данных сотрудников: {e}")
                raise e

        # Получаем роли привязанные к выбранному отделу по коду отдела
        current_roles = get_current_roles_by_department_code(department_code)
        # Заполняем список ролей
        for role in current_roles:
            item = QListWidgetItem(role.name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            item.setData(Qt.ItemDataRole.UserRole, {
                "role_id": role.id,
            })
            self.ui.roles_list.addItem(item)
        self.ui.roles_list.setCurrentRow(0)  # По умолчанию выбираем первый элемент списка

    def add_role(self) -> None:
        """
        Функция колбэк для кнопки добавления роли(Программы мотивации).
        Добавляет новую роль если такой нет.
        """
        try:
            current_department_code = get_department_code(self.ui.department)
        except Exception as e:
            print(f"[add_role] Ошибка при получении кода отдела {e}")
            # Если почему-то не выбран отдел выбрасываем сообщение
            confirm_dialog = QMessageBox(self)
            confirm_dialog.setWindowTitle("Ошибка, не выбран отдел")
            confirm_dialog.setText("Выберите отдел")
            confirm_dialog.setStandardButtons(QMessageBox.StandardButton.Ok)
            confirm_dialog.exec()
            return

        new_role_name = self.ui.roles_add_line.text()
        if not new_role_name:
            return

        with get_session() as session:
            # Получаем роли привязанные к выбранному отделу по коду отдела
            current_roles = get_current_roles_by_department_code(current_department_code)

            # Если на данном отделе еще нет роли с таким названием
            if new_role_name not in [role.name for role in current_roles]:
                motivation_program = MotivationProgram(name=new_role_name, department_code=current_department_code)
                session.add(motivation_program)
                session.commit()
            else:
                confirm_dialog = QMessageBox(self)
                confirm_dialog.setWindowTitle("Ошибка, роль уже существует")
                confirm_dialog.setText(f"Программа с именем '{new_role_name}' уже существует")
                confirm_dialog.setStandardButtons(QMessageBox.StandardButton.Ok)
                confirm_dialog.exec()
            self.ui.roles_add_line.clear()
            self.ui.roles_add_line.setFocus()
            self.set_current_roles()

    def delete_role(self) -> None:
        current_role = self.ui.roles_list.currentItem()

        if current_role:
            current_role_data = current_role.data(Qt.ItemDataRole.UserRole)
            role_id = current_role_data["role_id"]
            item_text = current_role.text()

            # Создание диалога подтверждения
            confirm_dialog = QMessageBox(self)
            confirm_dialog.setWindowTitle("Подтверждение")
            confirm_dialog.setText(f"Вы уверены, что хотите удалить программу'{item_text}'?")
            confirm_dialog.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            button_yes = confirm_dialog.button(QMessageBox.StandardButton.Yes)
            button_no = confirm_dialog.button(QMessageBox.StandardButton.No)
            button_yes.setText("Удалить")
            button_no.setText("Отмена")

            # Установка стиля
            confirm_dialog.setStyleSheet(CONFIRM_DIALOG_STYLE)

            reply = confirm_dialog.exec()

            if reply == QMessageBox.StandardButton.Yes:
                # Выполняем удаление
                delete_motivation_program(role_id)
                # Обновляем список ролей(Программ мотивации) в UI
                self.set_current_roles()

                print(f"Удаление {item_text} выполнено!")
            else:
                print("Удаление отменено!")

        else:
            msg_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Предупреждение",
                "Выберите программу для удаления!",
                QMessageBox.StandardButton.Ok,
                self
            )

            msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
            msg_box.exec()

    def update_role_in_db(self, item):
        """
        Функция колбэк для сохранения изменений роли (Программы мотивации).
        Сохраняет изменения роли в БД.
        """
        new_name = item.text()
        current_role_id = item.data(Qt.ItemDataRole.UserRole)["role_id"]

        with get_session() as session:
            current_roles = get_current_roles_by_department_code(self.ui.department.currentData()["department_code"])
            if new_name not in [role.name for role in current_roles]:
                session.query(MotivationProgram).filter(MotivationProgram.id == current_role_id
                                                        ).update({"name": new_name})
                session.commit()
                print(f"Изменение названия роли на '{new_name}' выполнено!")
                # Перерисовываем роли из БД, чтобы отразить актуальное состояние
            else:
                msg_box = QMessageBox(
                    QMessageBox.Icon.Warning,
                    "Предупреждение",
                    "Программа с таким именем уже существует!",
                    QMessageBox.StandardButton.Ok,
                    self
                )
                msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
                msg_box.exec()
            self.set_current_roles()

    def add_row(self):
        """
        Функция колбэк для кнопки добавления строки в таблицу настроек мотивации.
        Добавляет новую строку в таблицу.
        """
        # Если выбрана программа добавляем строку
        if current_role := self.ui.roles_list.currentItem():
            current_role_data: dict = current_role.data(Qt.ItemDataRole.UserRole)
            current_role_id: str = str(current_role_data["role_id"])

            current_row_count = self.ui.table_motivate_settings.rowCount()
            self.ui.table_motivate_settings.insertRow(current_row_count)
            # Добавляем в header информацию о роли к которой привязана таблица
            # (3 столбец скрыт и используется длля хнанения роли)
            self.ui.table_motivate_settings.horizontalHeaderItem(2).setText(current_role_id)

        else:
            msg_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Предупреждение",
                "Пожалуйста, выберите программу к которой хотите добавить мотивацию.",
                QMessageBox.StandardButton.Ok,
                self
            )

            msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
            msg_box.exec()

    def remove_selected_row(self):
        """
        Функция колбэк для кнопки удаления строки в таблице настроек мотивации.
        Удаляет выбранную строку в таблице.
        """
        selected_row = self.ui.table_motivate_settings.currentRow()
        if selected_row >= 0:
            self.ui.table_motivate_settings.removeRow(selected_row)
            self.threshold_table_controller.changes_made = True
        else:
            msg_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Предупреждение",
                "Пожалуйста, выберите строку для удаления.",
                QMessageBox.StandardButton.Ok,
                self
            )

            msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
            msg_box.exec()

    def fill_role_settings_table(self):
        """
        Функция заполнения таблицы настроек мотивации.
        Заполняет таблицу настройками мотивации для выбранной роли.
        """
        # Проверяем есть ли несохраненные изменения
        if self.threshold_table_controller.changes_made:
            msg_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Предупреждение",
                f"Есть не сохраненные изменения программы мотивации.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                self
            )

            button_yes = msg_box.button(QMessageBox.StandardButton.Yes)
            button_no = msg_box.button(QMessageBox.StandardButton.No)
            #button_cancel = msg_box.button(QMessageBox.StandardButton.Cancel)
            button_yes.setText("Сохранить")
            button_no.setText("Отменить")
            #button_cancel.setText("Вернуться к редактированию программы мотивации")

            msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
            reply = msg_box.exec()
            if reply == QMessageBox.StandardButton.Yes:
                self.save_table_data_to_db(on_save_button_push=False)  # Даём понять что сохранение не по кнопке
            elif reply == QMessageBox.StandardButton.No:
                self.threshold_table_controller.changes_made = False

        # Блокируем сигналы чтобы не перехватывать изменения в таблице
        self.ui.table_motivate_settings.blockSignals(True)

        # Отчищаем таблицу
        # TODO: Перенести
        self.ui.table_motivate_settings.setRowCount(0)
        self.ui.employees_table.setRowCount(0)

        if current_role := self.ui.roles_list.currentItem():
            print(f"Current role :{current_role}")
            current_role_data: dict = current_role.data(Qt.ItemDataRole.UserRole)
            current_role_id: str = str(current_role_data["role_id"])

            # Заполняем список сотрудников привязанных к роли
            self.fill_employees_table(current_role_id)
        else:
            print(f"Current role не указана, оставляем таблицу пустой")
            # Разблокируем сигналы
            self.ui.table_motivate_settings.blockSignals(False)
            return
        with get_session() as session:

            self.ui.table_motivate_settings.horizontalHeaderItem(2).setText(current_role_id)

            current_motivation_program: MotivationProgram = session.query(MotivationProgram).filter(
                MotivationProgram.id == current_role_id).first()

            # Загружаем данные в таблицу из db используя контроллер
            self.threshold_table_controller.load_data(current_motivation_program)

            # Разблокируем сигналы
            self.ui.table_motivate_settings.blockSignals(False)

    def save_table_data_to_db(self, on_save_button_push: bool = False):
        """
        Сохраняет данные таблицы в базу данных.

        Args:
            on_save_button_push (bool): Флаг, указывающий, было ли сохранение инициировано нажатием кнопки
        """
        try:
            # Получаем id программы мотивации из заголовка таблицы
            current_table_role_id = self.ui.table_motivate_settings.horizontalHeaderItem(2).text()

            # Проверяем, привязана ли таблица к программе мотивации
            if current_table_role_id == 'no_role':
                QMessageBox.warning(
                    self,
                    "Предупреждение",
                    "Таблица пуста и не привязана ни к одной программе мотивации",
                    QMessageBox.StandardButton.Ok
                )
                return

            with get_session() as session:
                # Получаем программу мотивации
                current_motivation_program = session.query(MotivationProgram).filter(
                    MotivationProgram.id == current_table_role_id
                ).first()

                # Если программа не найдена
                if not current_motivation_program:
                    raise ValueError("Программа мотивации не найдена")

                # Сохраняем данные используя контроллер таблицы
                if self.threshold_table_controller.save_data(current_motivation_program, session):
                    self.threshold_table_controller.changes_made = False

                    # Если сохранение было по кнопке, обновляем отображение таблицы
                    if on_save_button_push:
                        self.fill_role_settings_table()

                    msg_box = QMessageBox(
                        QMessageBox.Icon.Information,
                        "Успех",
                        "Данные успешно сохранены",
                        QMessageBox.StandardButton.Ok,
                        self
                    )

                    msg_box.setStyleSheet(WARNING_DIALOG_STYLE)  # TODO: Сделать шаблон  SUCCESS_DIALOG_STYLE
                    msg_box.exec()

        except Exception as e:
            msg_box = QMessageBox(
                QMessageBox.Icon.Critical,
                "Ошибка",
                f"Не удалось сохранить данные: {str(e)}",
                QMessageBox.StandardButton.Ok,
                self
            )

            msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
            msg_box.exec()

    def fill_employees_table(self, current_role_id):
        """
        Функция заполнения таблицы сотрудников.
        Заполняет таблицу сотрудниками для выбранной роли.
        """
        # Обновляем данные по работникам во внутренней БД
        with get_session() as session:
            employees: list[Type[Employee]] = get_employees_by_motivation_program_id(
                session=session,
                motivation_program_id=current_role_id
            )

            self.ui.employees_table.setRowCount(0)
            self.ui.employees_table.setRowCount(len(employees))

            for i, employee in enumerate(employees):
                self.ui.employees_table.setItem(i, 0, QTableWidgetItem(employee.name))
                self.ui.employees_table.setItem(i, 1, QTableWidgetItem(employee.position))
                self.ui.employees_table.setItem(i, 3, QTableWidgetItem(str(employee.code)))

                departments_string = ""
                for department in employee.departments:
                    departments_string += department.name + " "
                self.ui.employees_table.setItem(i, 2, QTableWidgetItem(departments_string))

    def open_edit_employees_window(self):
        """
        Открывает окно добавления, удаления сотрудников из программы мотивации
        :return:
        """
        motivation_program = self.ui.roles_list.currentItem()
        print(f"Мотивационная программа в функции открытия окна получена - {motivation_program}")
        if not motivation_program:
            msg_box = QMessageBox(
                QMessageBox.Icon.Warning,
                "Предупреждение",
                "Не выбрана программа мотивации",
                QMessageBox.StandardButton.Ok,
                self
            )

            msg_box.setStyleSheet(WARNING_DIALOG_STYLE)
            msg_box.exec()

            return

        motivation_program_data = motivation_program.data(Qt.ItemDataRole.UserRole)
        motivation_program_id = str(motivation_program_data["role_id"])
        motivation_program_name = motivation_program.text()
        print(f"Передаю role_id в новое окно - {motivation_program_id}")
        # Сохраняем ссылку на окно, чтобы оно сразу не закрылось сборщиком мусора.
        self.edit_employees_window = EditEmployeesWindow(motivation_program_id, parent=self)
        self.edit_employees_window.setWindowTitle(f"Сотрудники программы мотивации {motivation_program_name}")
        self.edit_employees_window.show()
        self.edit_employees_window.exec()
        self.fill_employees_table(motivation_program_id)
        print("Закрыто")

    def export_to_excel(self):
        """
        Экспортирует данные из сводной таблицы зарплат в Excel
        """
        # Создаем новый документ
        wb = Workbook()
        ws = wb.active
        ws.Name = "Сводная таблица зарплат"
        ws.cell(1, 1).value = "ФИО"
        ws.cell(1, 2).value = "Полное имя"
        ws.cell(1, 3).value = "Сумма зарплаты\nc 1 по 15"
        ws.cell(1, 4).value = "Кол-во полных смен\nс 1 по 15"
        ws.cell(1, 5).value = "Кол-во неполных смен\nс 1 по 15"
        ws.cell(1, 6).value = "Сумма зарплаты\nc 16-го до конца месяца"
        ws.cell(1, 7).value = "Кол-во полных смен\nc 16-го до конца месяца"
        ws.cell(1, 8).value = "Кол-во неполных смен\nc 16-го до конца месяца"
        ws.cell(1, 9).value = "Сумма зарплаты\nза весь месяц"
        ws.cell(1, 10).value = "Кол-во полных смен\nза весь месяц"
        ws.cell(1, 11).value = "Кол-во неполных смен\nза весь месяц"
        ws.cell(1, 12).value = "Должность"
        ws.cell(1, 13).value = "Отдел"
        ws.cell(1, 14).value = "Табельный номер"
        ws.cell(1, 15).value = "id"

        # Устанавливаем выравнивание по центру и перенос текста, шрифт, бордеры, а также заливку
        for col in range(1, 16):
            ws.cell(1, col).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            ws.cell(1, col).fill = PatternFill(
                start_color="00339966",
                end_color="00339966",
                fill_type="solid",
            )

            ws.cell(1, col).border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

            ws.cell(1, col).font = Font(
                bold=True,
                color="00FFFFFF",
                name="Arial",
            )

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 23
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 23
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 20

        for row in range(self.ui.salar_table.rowCount()):
            for col in range(self.ui.salar_table.columnCount()):
                item = self.ui.salar_table.item(row, col)
                ws.cell(row=row + 2, column=col + 1).value = item.text()

        wb.save("salary_table.xlsx")

        os.startfile("salary_table.xlsx")

    def payslip_report_callback(self):
        """
        Функция, которая вызывается при нажатии на кнопку "Отчет по зарплате"
        """
        logger.info("=========================Отчет по зарплате=========================")
        try:
            self.payslip_generator.create_payslip_pdf(
                self.ui.date_from.date().toPython(), self.ui.date_to.date().toPython()
            )
        except PermissionError as e:
            self.show_error_message("Отчет открыт в другой программе. Закройте другие программы использующие отчет"
                                    " и повторите попытку.")
        except Exception as e:
            self.show_error_message(f"Ошибка при создании отчета: Непредвиденная ошибка: {e}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = SalaryReader()
    window.show()
    sys.exit(app.exec())
